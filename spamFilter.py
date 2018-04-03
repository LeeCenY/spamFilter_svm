import openpyxl
import numpy as np
from cleanText import cleanString
from collections import Counter
from sklearn.svm import SVC, NuSVC, LinearSVC
from sklearn.naive_bayes import MultinomialNB, GaussianNB, BernoulliNB
from sklearn.metrics import confusion_matrix
from sklearn.cross_validation import train_test_split

# Get the original dataset
def store():

    workBookOld = openpyxl.load_workbook('DataSet.xlsx')
    dataSheetOld = workBookOld['Data set']

    xData = []
    yData = []

    rows = dataSheetOld.max_row

    for i in range(2, rows+1):

        if (str(dataSheetOld.cell(row = i+2, column = 2).value) != 'None'):
            xData.append(cleanString(str(dataSheetOld.cell(row = i+2, column = 1).value)))
            if (str(dataSheetOld.cell(row = i+2, column = 2).value) == "Spam"):
                yData.append(1)
            else:
                yData.append(0)

    # NOTE: to train data on the entire dataset, simply return xData and yData
    # Splitting the data like this is to obtain test cases and calculate the F-score of the learning algorithm
    xTrain, xTest, yTrain, yTest = train_test_split(xData, yData, test_size = 0.2, random_state = 42)
    return xTrain, xTest, yTrain, yTest

# make a dictionary of the most common words
def makeDictionary(xData):

    all_words = []
    
    for mail in xData:

        words = mail.split()
        all_words.extend(words)
    
    dictionary = Counter(all_words)
    dictionary = dictionary.most_common(3000)
    return dictionary

# construct a feature vector for each mail
def extractFeatures(xData, dictionary):
    
    featureMatrix = np.zeros((len(xData), 3000))
    emailId = 0

    for mail in xData:
        for word in mail:
            wordId = 0
            for i,d in enumerate(dictionary):
                if (d[0] == word):
                    wordId = i
                    featureMatrix[emailId, wordId] = mail.count(word)
        emailId += 1

    return featureMatrix

# Create training data
xTrain, xTest, yTrain, yTest = store()
trainDictionary = makeDictionary(xTrain)

# Create feature vector and matrix for yTrain and xTrain
yTrainMatrix = np.zeros(len(yTrain))
for i in range(len(yTrain)):
    if (yTrain[i] == 1):
        yTrainMatrix[i] = 1

xTrainMatrix = extractFeatures(xTrain, trainDictionary)

# Training SVM classifier
model = NuSVC(nu = 0.01, class_weight = 'balanced')
model.fit(xTrainMatrix, yTrainMatrix)

# Calculating the F-score
def calcFScore(xTest, yTest):

    xTestMatrix = extractFeatures(xTest, trainDictionary)
    yTestMatrix = np.zeros(len(yTest))
    for i in range(len(yTest)):
        if (yTest[i] == 1):
            yTestMatrix[i] = 1

    result = model.predict(xTestMatrix)
    matrix = confusion_matrix(result, yTestMatrix)

    precision = float(matrix[0][0])/(matrix[0][0] + matrix[0][1])
    recall = float(matrix[0][0])/(matrix[0][0] + matrix[1][0])

    fScore = (2 * precision * recall)/(precision + recall)
    return fScore, matrix

# Test new data for Spam
def predict(emailBody):

    featureMatrix = extractFeatures([emailBody], trainDictionary)
    result = model.predict(featureMatrix)

    if (1 in result):
        return "Spam"
    else:
        return "Not Spam"